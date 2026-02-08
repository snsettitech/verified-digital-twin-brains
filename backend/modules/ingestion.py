import os
import uuid
import asyncio
import re
import json
import feedparser
import yt_dlp
import time
import httpx
import html
import html as html_lib
import asyncio
from typing import List, Dict, Optional, Any, Tuple
from bs4 import BeautifulSoup
from modules.transcription import transcribe_audio_multi
from modules.embeddings import get_embedding
from PyPDF2 import PdfReader
import docx
import openpyxl
from youtube_transcript_api import YouTubeTranscriptApi
from modules.clients import get_openai_client, get_pinecone_index
from modules.observability import supabase, log_ingestion_event
from modules.ingestion_diagnostics import start_step, finish_step, build_error
from modules.health_checks import run_all_health_checks, calculate_content_hash
from modules.access_groups import get_default_group, add_content_permission
from modules.governance import AuditLogger


# ============================================================================
# Enterprise Configuration & Utilities
# ============================================================================

class YouTubeConfig:
    """Enterprise-grade YouTube ingestion configuration."""
    MAX_RETRIES = int(os.getenv("YOUTUBE_MAX_RETRIES", "5"))
    ASR_MODEL = os.getenv("YOUTUBE_ASR_MODEL", "whisper-large-v3")
    ASR_PROVIDER = os.getenv("YOUTUBE_ASR_PROVIDER", "openai")
    LANGUAGE_DETECTION = os.getenv("YOUTUBE_LANGUAGE_DETECTION", "true").lower() == "true"
    PII_SCRUB = os.getenv("YOUTUBE_PII_SCRUB", "true").lower() == "true"
    VERBOSE_LOGGING = os.getenv("YOUTUBE_VERBOSE_LOGGING", "false").lower() == "true"
    COOKIES_FILE = os.getenv("YOUTUBE_COOKIES_FILE")
    PROXY = os.getenv("YOUTUBE_PROXY")


class ErrorClassifier:
    """Classify YouTube/yt-dlp errors for better handling and telemetry."""
    
    @staticmethod
    def classify(error_msg: str) -> Tuple[str, str, bool]:
        """
        Classify error and return (category, user_message, is_retryable).
        
        Categories:
        - auth: Authentication/login required
        - rate_limit: HTTP 429, quota exceeded
        - gating: Age/region/access restrictions
        - unavailable: Video deleted/private
        - network: Connection/timeout issues
        - unknown: Unclassified
        """
        error_lower = error_msg.lower()
        
        # Rate limiting
        if "429" in error_msg or "rate" in error_lower or "quota" in error_lower or "too many requests" in error_lower:
            return "rate_limit", "YouTube rate limit reached. Retrying with backoff...", True
        
        # Authentication required
        if "403" in error_msg or "sign in" in error_lower or "unauthorized" in error_lower:
            return "auth", "This video requires authentication or is age-restricted.", False
        
        # Gating (region, age, etc.)
        if "geo" in error_lower or "region" in error_lower or "not available" in error_lower:
            return "gating", "This video is not available in your region.", False
        
        # Video unavailable
        if "unavailable" in error_lower or "deleted" in error_lower or "not found" in error_lower:
            return "unavailable", "This video is unavailable (deleted, private, or not found).", False
        
        # Network issues
        if "timeout" in error_lower or "connection" in error_lower or "socket" in error_lower:
            return "network", "Network connection issue. Retrying...", True
        
        return "unknown", f"Unexpected error: {error_msg}", False


class LanguageDetector:
    """Lightweight language detection for transcripts."""
    
    # Simple heuristics for common languages (word counts and patterns)
    LANGUAGE_PATTERNS = {
        "en": [r"\b(the|and|a|to|of|in|is|that|it)\b"],
        "es": [r"\b(el|la|de|que|y|a|en|un|es)\b"],
        "fr": [r"\b(le|la|de|et|a|en|un|est|c'est)\b"],
        "de": [r"\b(der|die|und|in|den|von|zu|das)\b"],
        "ja": [r"[ぁ-ん]|[ァ-ヴー]|[一-龥]"],
        "zh": [r"[\u4e00-\u9fff]"],
    }
    
    @staticmethod
    def detect(text: str) -> str:
        """Detect language of text. Returns language code (e.g., 'en', 'es')."""
        if not text or len(text) < 10:
            return "en"  # Default
        
        text_lower = text.lower()
        scores = {}
        
        for lang, patterns in LanguageDetector.LANGUAGE_PATTERNS.items():
            count = 0
            for pattern in patterns:
                count += len(re.findall(pattern, text_lower))
            scores[lang] = count
        
        if scores:
            return max(scores, key=scores.get)
        return "en"


class PIIScrubber:
    """Lightweight PII detection and redaction."""
    
    # Regex patterns for common PII (very permissive, for flagging)
    PATTERNS = {
        "email": r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}",
        "phone": r"(?:\+1[-.\s]?)?\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})\b",  # Matches multiple formats
        "ssn": r"\b\d{3}-\d{2}-\d{4}\b",
        "credit_card": r"\b\d{4}[\s-]?\d{4}[\s-]?\d{4}[\s-]?\d{4}\b",
        "ip_address": r"\b(?:\d{1,3}\.){3}\d{1,3}\b",
    }
    
    @staticmethod
    def detect_pii(text: str) -> Dict[str, List[str]]:
        """Detect PII in text. Returns dict of pii_type -> [matches]."""
        detected = {}
        for pii_type, pattern in PIIScrubber.PATTERNS.items():
            matches = re.findall(pattern, text)
            if matches:
                detected[pii_type] = matches
        return detected
    
    @staticmethod
    def has_pii(text: str) -> bool:
        """Check if text contains any PII."""
        return bool(PIIScrubber.detect_pii(text))
    
    @staticmethod
    def scrub(text: str) -> str:
        """Redact PII from text."""
        for pii_type, pattern in PIIScrubber.PATTERNS.items():
            placeholder = f"[{pii_type.upper()}]"
            text = re.sub(pattern, placeholder, text)
        return text


def extract_text_from_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    text_parts: List[str] = []
    for page in reader.pages:
        # Some PDFs return None for image-only pages; keep extraction best-effort
        # instead of crashing the whole ingest on a single page.
        page_text = page.extract_text() or ""
        if page_text:
            text_parts.append(page_text)
    return "\n".join(text_parts)


def extract_text_from_docx(file_path: str) -> str:
    """Extract text from a Word document."""
    doc = docx.Document(file_path)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return "\n".join(text)


def extract_text_from_excel(file_path: str) -> str:
    """Extract text from all sheets in an Excel file."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text.append(f"--- Sheet: {sheet_name} ---")
        for row in sheet.iter_rows(values_only=True):
            # Filter out None values and convert to string
            row_text = [str(cell) for cell in row if cell is not None]
            if row_text:
                text.append(" | ".join(row_text))
    return "\n".join(text)


def extract_video_id(url: str) -> str:
    """
    Extracts the video ID from a YouTube URL.
    """
    patterns = [
        r'(?:v=|\/)([0-9A-Za-z_-]{11}).*',
        r'(?:embed\/)([0-9A-Za-z_-]{11}).*',
        r'(?:youtu\.be\/)([0-9A-Za-z_-]{11}).*'
    ]
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None


def detect_url_provider(url: str) -> str:
    u = (url or "").strip().lower()
    if "youtube.com" in u or "youtu.be" in u:
        return "youtube"
    if "twitter.com" in u or "x.com" in u:
        return "x"
    if "linkedin.com" in u:
        return "linkedin"
    if u.endswith(".rss") or "feed" in u or "podcast" in u or "anchor.fm" in u or "podbean" in u:
        return "podcast"
    if u.startswith("http://") or u.startswith("https://"):
        return "web"
    return "unknown"




def _build_yt_dlp_opts(video_id: str, source_id: str, twin_id: str) -> dict:
    """
    Build yt-dlp options with enterprise-grade toggles:
    - Multiple client emulation strategies (Android, web, iOS)
    - Optional cookies (file only) for age/region-gated videos
    - Optional proxy for IP reputation routing
    - Browser headers to avoid bot detection
    """
    temp_dir = "temp_uploads"
    os.makedirs(temp_dir, exist_ok=True)
    temp_filename = f"yt_{video_id}"

    cookie_file = os.getenv("YOUTUBE_COOKIES_FILE")
    proxy_url = os.getenv("YOUTUBE_PROXY")

    # Start with Android client emulation (most reliable for containers)
    ydl_opts = {
        'format': 'm4a/bestaudio/best',
        'outtmpl': os.path.join(temp_dir, f"{temp_filename}.%(ext)s"),
        'postprocessors': [{
            'key': 'FFmpegExtractAudio',
            'preferredcodec': 'mp3',
            'preferredquality': '192',
        }],
        'quiet': True,
        'no_warnings': True,
        # Multiple client strategies (try Android first, then web/iOS)
        'extractor_args': {
            'youtube': {
                'player_client': ['android', 'web', 'ios'],
                'player_skip': ['webpage', 'configs', 'js'],
                'include_live_dash': [True],
            }
        },
        'nocheckcertificate': True,
        'socket_timeout': 30,
        'http_headers': {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept-Language': 'en-US,en;q=0.9',
        },
        'retries': 5,
        'fragment_retries': 5,
        'skip_unavailable_fragments': True,
    }

    if proxy_url:
        ydl_opts['proxy'] = proxy_url
        log_ingestion_event(source_id, twin_id, "info", "Using YOUTUBE_PROXY for YouTube download")

    # Only use file-based cookies (cookiesfrombrowser doesn't work in containers)
    if cookie_file and os.path.exists(cookie_file):
        ydl_opts['cookiefile'] = cookie_file
        log_ingestion_event(source_id, twin_id, "info", "Using YOUTUBE_COOKIES_FILE for YouTube download")
    else:
        if YouTubeConfig.VERBOSE_LOGGING:
            log_ingestion_event(source_id, twin_id, "info", "No cookie file found; using client emulation only")

    # Log configuration for telemetry
    log_ingestion_event(source_id, twin_id, "info", 
        f"YouTube config: ASR={YouTubeConfig.ASR_PROVIDER}:{YouTubeConfig.ASR_MODEL}, "
        f"MaxRetries={YouTubeConfig.MAX_RETRIES}, LanguageDetection={YouTubeConfig.LANGUAGE_DETECTION}, "
        f"PIIScrub={YouTubeConfig.PII_SCRUB}")

    return ydl_opts, temp_dir, temp_filename


def _transcript_items_to_text(items: Any) -> Optional[str]:
    """Normalize transcript payloads from different youtube-transcript-api versions."""
    if items is None:
        return None

    raw_items = items
    if hasattr(items, "to_raw_data"):
        raw_items = items.to_raw_data()

    if raw_items is None:
        return None

    if not isinstance(raw_items, list):
        try:
            raw_items = list(raw_items)
        except Exception:
            return None

    parts: List[str] = []
    for item in raw_items:
        if isinstance(item, dict):
            parts.append(str(item.get("text", "")))
        elif isinstance(item, str):
            parts.append(item)

    text = " ".join(parts).strip()
    return text or None


def fetch_youtube_transcript_compat(video_id: str, yta_cls: Optional[Any] = None) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Fetch transcript text while supporting both old and new youtube-transcript-api APIs.

    Returns:
      (text, error_message, method_used)
      method_used is one of: fetch, list, get_transcript, list_transcripts.
    """
    transcript_error: Optional[str] = None

    try:
        if yta_cls is None:
            from youtube_transcript_api import YouTubeTranscriptApi as yta_cls  # type: ignore[no-redef]

        # New API: instance.fetch(video_id, languages=...)
        try:
            api = yta_cls()
            if hasattr(api, "fetch"):
                fetched = api.fetch(video_id, languages=("en", "en-US", "en-GB"))
                text = _transcript_items_to_text(fetched)
                if text:
                    return text, None, "fetch"
        except Exception as e:
            transcript_error = str(e)

        # New API fallback: instance.list(video_id).find_*_transcript(...).fetch()
        try:
            api = yta_cls()
            if hasattr(api, "list"):
                transcript_list = api.list(video_id)
                transcript = None
                try:
                    if hasattr(transcript_list, "find_manually_created_transcript"):
                        transcript = transcript_list.find_manually_created_transcript(["en", "en-US", "en-GB"])
                except Exception:
                    transcript = None
                if transcript is None:
                    try:
                        if hasattr(transcript_list, "find_generated_transcript"):
                            transcript = transcript_list.find_generated_transcript(["en", "en-US", "en-GB"])
                    except Exception:
                        transcript = None
                if transcript is None:
                    transcript = transcript_list.find_transcript(["en", "en-US", "en-GB"])

                fetched = transcript.fetch(preserve_formatting=False)
                text = _transcript_items_to_text(fetched)
                if text:
                    return text, None, "list"
        except Exception as e:
            if not transcript_error:
                transcript_error = str(e)

        # Old API fallback: static get_transcript(video_id)
        if hasattr(yta_cls, "get_transcript"):
            try:
                items = yta_cls.get_transcript(video_id)
                text = _transcript_items_to_text(items)
                if text:
                    return text, None, "get_transcript"
            except Exception as e:
                if not transcript_error:
                    transcript_error = str(e)

        # Old API fallback: static list_transcripts(video_id)
        if hasattr(yta_cls, "list_transcripts"):
            try:
                transcript_list = yta_cls.list_transcripts(video_id)

                for transcript in getattr(transcript_list, "manually_created_transcripts", []) or []:
                    try:
                        text = _transcript_items_to_text(transcript.fetch())
                        if text:
                            return text, None, "list_transcripts"
                    except Exception:
                        continue

                for transcript in getattr(transcript_list, "generated_transcripts", []) or []:
                    try:
                        text = _transcript_items_to_text(transcript.fetch())
                        if text:
                            return text, None, "list_transcripts"
                    except Exception:
                        continue
            except Exception as e:
                if not transcript_error:
                    transcript_error = str(e)
    except Exception as e:
        transcript_error = str(e)

    return None, transcript_error, None


async def ingest_youtube_transcript(source_id: str, twin_id: str, url: str, correlation_id: Optional[str] = None):
    """
    Ingest YouTube video transcript using robust multi-strategy approach.
    """
    video_id = extract_video_id(url)
    if not video_id:
        raise ValueError("Invalid YouTube URL")

    provider = "youtube"
    # -------------------------------------------------------------
    # Step 0: Ensure Source Record exists
    # -------------------------------------------------------------
    try:
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": f"YouTube: {video_id}",
            "file_size": 0,
            "content_text": "",
            "status": "processing",
            "staging_status": "staged",
            "citation_url": url,
        }).execute()
    except Exception as e:
        # Don't crash ingestion for telemetry upsert issues; later updates may still succeed.
        print(f"[YouTube] Warning: Failed to upsert source record: {e}")

    text = None
    video_title = None
    transcript_error = None
    fetch_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        correlation_id=correlation_id,
        message="Fetching YouTube transcript/captions",
        metadata={"video_id": video_id, "url": url},
    )

    # -------------------------------------------------------------
    # Step 1: Validate via YouTube Data API (if key available)
    # -------------------------------------------------------------
    google_api_key = os.getenv("GOOGLE_API_KEY")
    if google_api_key:
        try:
            from googleapiclient.discovery import build
            youtube = build('youtube', 'v3', developerKey=google_api_key)
            request = youtube.videos().list(part="snippet", id=video_id)
            response = request.execute()

            if response["items"]:
                snippet = response["items"][0]["snippet"]
                video_title = snippet["title"]
                # Update filename with real title
                supabase.table("sources").update({
                    "filename": f"YouTube: {video_title}"
                }).eq("id", source_id).execute()
                print(f"[YouTube] Validated video: {video_title}")
        except Exception as e:
            print(f"[YouTube] Data API check failed (non-blocking): {e}")

    # -------------------------------------------------------------
    # Strategy 1: youtube-transcript-api (Fast, No Download)
    # -------------------------------------------------------------
    text, transcript_error, transcript_method = fetch_youtube_transcript_compat(video_id)
    if text:
        log_ingestion_event(
            source_id,
            twin_id,
            "info",
            f"Fetched YouTube transcript via youtube-transcript-api ({transcript_method})"
        )
    elif transcript_error and YouTubeConfig.VERBOSE_LOGGING:
        print(f"[YouTube] Transcript API failed: {transcript_error}")

    # -------------------------------------------------------------
    # Strategy 1.6: Direct HTTP fetch with browser headers (bypasses library blocks)
    # -------------------------------------------------------------
    if not text:
        try:
            print(f"[YouTube] Trying direct HTTP transcript fetch for {video_id}")
            log_ingestion_event(source_id, twin_id, "info", "Attempting direct HTTP transcript fetch")
            
            # Use Chrome-like headers to appear as a real browser
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
            }
            
            # Fetch the video page
            video_url = f"https://www.youtube.com/watch?v={video_id}"
            async with httpx.AsyncClient(timeout=30, headers=headers, follow_redirects=True) as client:
                response = await client.get(video_url)
                
                if response.status_code == 200:
                    page_content = response.text
                    
                    # Extract caption track URLs from the page
                    # Strategy 1.6.a: Main Player Response
                    caption_match = re.search(r'"captionTracks":\s*(\[.*?\])', page_content)
                    
                    # Strategy 1.6.b: Escaped JSON (mobile/older formats)
                    if not caption_match:
                        caption_match = re.search(r'captionTracks\\":(\[.*?\])', page_content)
                    
                    # Strategy 1.6.c: Alternative escaped format
                    if not caption_match:
                        caption_match = re.search(r'\\u0022captionTracks\\u0022:\\s*(\\\[.*?\\\])', page_content)

                    if caption_match:
                        print(f"[YouTube] Found caption tracks on page")
                        try:
                            raw_json = caption_match.group(1).replace('\\"', '"').replace('\\\\', '\\').replace('\\u0022', '"')
                            # Ensure we handle double brackets from some escaped formats
                            if raw_json.startswith('\\['): raw_json = json.loads(f'"{raw_json}"')
                            caption_tracks = json.loads(raw_json)
                            
                            # Find English or first available caption
                            caption_url = None
                            for track in caption_tracks:
                                lang = track.get("languageCode", "")
                                if lang.startswith("en"):
                                    caption_url = track.get("baseUrl")
                                    break
                            
                            # Fallback to first track if no English
                            if not caption_url and caption_tracks:
                                caption_url = caption_tracks[0].get("baseUrl")
                            
                            if caption_url:
                                # Fetch the actual captions
                                caption_response = await client.get(caption_url)
                                if caption_response.status_code == 200:
                                    # Parse XML captions
                                    caption_xml = caption_response.text
                                    caption_texts = re.findall(r'<text[^>]*>(.*?)</text>', caption_xml, re.DOTALL)
                                    
                                    if caption_texts:
                                        # Unescape HTML entities and join
                                        text = " ".join([html.unescape(t.strip()) for t in caption_texts])
                                        text = re.sub(r'\s+', ' ', text).strip()
                                        log_ingestion_event(source_id, twin_id, "info", f"Direct HTTP transcript fetch successful ({len(text)} chars)")
                                        print(f"[YouTube] Direct HTTP fetch succeeded: {len(text)} characters")
                        except json.JSONDecodeError:
                            print(f"[YouTube] Could not parse caption tracks JSON")
        except Exception as e:
            print(f"[YouTube] Direct HTTP fetch failed: {e}")

    # -------------------------------------------------------------
    # Strategy 2: yt-dlp with Multiple Client Emulation & Better Error Handling
    # Using YouTubeRetryStrategy for enterprise-grade retry logic
    # -------------------------------------------------------------
    if not text:
        print("[YouTube] No captions found. Starting robust audio download...")
        log_ingestion_event(source_id, twin_id, "warning", "Attempting robust audio download (yt-dlp with multiple clients)")

        from modules.youtube_retry_strategy import YouTubeRetryStrategy
        
        ydl_opts, temp_dir, temp_filename = _build_yt_dlp_opts(video_id, source_id, twin_id)
        
        # Initialize retry strategy with configured max_retries
        config = YouTubeConfig()
        strategy = YouTubeRetryStrategy(
            source_id=source_id,
            twin_id=twin_id,
            max_retries=config.MAX_RETRIES,
            verbose=config.VERBOSE_LOGGING
        )
        
        temp_audio_path = None
        
        while strategy.attempts < strategy.max_retries and not text:
            attempt_no = strategy.attempts + 1
            try:
                print(f"[YouTube] Download attempt {attempt_no}/{strategy.max_retries} for {video_id}")
                
                # Download audio
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    ydl.download([url])

                temp_audio_path = os.path.join(temp_dir, f"{temp_filename}.mp3")
                
                if not os.path.exists(temp_audio_path):
                    raise FileNotFoundError(f"Audio file not created at {temp_audio_path}")

                # Transcribe
                print(f"[YouTube] Transcribing audio for {video_id}")
                text = transcribe_audio_multi(temp_audio_path)
                
                # Language detection
                language = LanguageDetector.detect(text)
                
                # PII detection
                pii_detected = PIIScrubber.has_pii(text) if config.PII_SCRUB else False
                detected_pii = PIIScrubber.detect_pii(text) if pii_detected else []
                
                # Log success with metadata
                strategy.attempts = attempt_no
                strategy.log_success(
                    content_length=len(text),
                    metadata={
                        "language": language,
                        "has_pii": pii_detected,
                        "pii_count": len(detected_pii)
                    }
                )
                
                log_ingestion_event(
                    source_id, twin_id, "info",
                    f"Audio transcribed via Gemini/Whisper ({len(text)} chars, lang={language}, pii={pii_detected})"
                )
                
                print(f"[YouTube] Successfully transcribed {video_id}: {len(text)} characters, language={language}")
                
            except Exception as download_error:
                error_msg = str(download_error)
                error_category, user_msg, retryable = ErrorClassifier.classify(error_msg)
                
                strategy.log_attempt(error_msg)
                
                print(f"[YouTube] Attempt {strategy.attempts} failed [{error_category}]: {error_msg}")
                log_ingestion_event(
                    source_id, twin_id, "warning",
                    f"Download attempt {strategy.attempts} failed [{error_category}]: {error_msg}"
                )
                
                # Determine if we should retry
                if not strategy.should_retry(error_category):
                    print(f"[YouTube] Error category '{error_category}' is non-retryable, stopping attempts")
                    break
                
                if strategy.attempts < strategy.max_retries:
                    wait_time = strategy.wait_for_retry()
                    print(f"[YouTube] Waiting {wait_time}s before retry...")
                else:
                    break

        # Cleanup temp file
        if temp_audio_path and os.path.exists(temp_audio_path):
            try:
                os.remove(temp_audio_path)
            except:
                pass

        if not text:
            # Get final error message with classification
            final_error = strategy.get_final_error_message(strategy.last_error or "Unknown error")
            metrics = strategy.get_metrics()
            final_error_category, _, _ = ErrorClassifier.classify(strategy.last_error or final_error)
            
            log_ingestion_event(
                source_id, twin_id, "error",
                f"YouTube ingestion failed after {strategy.attempts} attempts. Metrics: {metrics}"
            )
            err = build_error(
                code="YOUTUBE_TRANSCRIPT_UNAVAILABLE",
                message=final_error,
                provider=provider,
                step="fetching",
                provider_error_code=final_error_category,
                correlation_id=correlation_id,
                raw={
                    "video_id": video_id,
                    "url": url,
                    "attempts": strategy.attempts,
                    "metrics": metrics,
                    "last_error": strategy.last_error,
                },
            )
            finish_step(
                event_id=fetch_event_id,
                source_id=source_id,
                twin_id=twin_id,
                provider=provider,
                step="fetching",
                status="error",
                correlation_id=correlation_id,
                error=err,
            )
            raise ValueError(final_error)

    if not text:
        err = build_error(
            code="YOUTUBE_TRANSCRIPT_UNAVAILABLE",
            message="No transcript could be extracted. This video may not have captions.",
            provider=provider,
            step="fetching",
            correlation_id=correlation_id,
            raw={"video_id": video_id, "url": url, "transcript_error": transcript_error},
        )
        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise ValueError(
            "No transcript could be extracted. This video may not have captions. "
            "Try a different video with closed captions (CC) enabled."
        )

    finish_step(
        event_id=fetch_event_id,
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        status="completed",
        correlation_id=correlation_id,
        metadata={"text_len": len(text), "has_title": bool(video_title)},
    )

    parsed_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="parsed",
        correlation_id=correlation_id,
        message="Persisting transcript to sources.content_text",
    )

    try:
        # Direct indexing - extract and index immediately
        content_hash = calculate_content_hash(text)

        # Update existing source row (do NOT insert again; the row is created when job is queued)
        filename = f"YouTube: {video_title or video_id}"
        supabase.table("sources").update({
            "filename": filename,
            "file_size": len(text),
            "content_text": text,
            "content_hash": content_hash,
            "status": "processing",
            "staging_status": "staged",
            "extracted_text_length": len(text),
            "citation_url": url,
        }).eq("id", source_id).eq("twin_id", twin_id).execute()

        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="completed",
            correlation_id=correlation_id,
            metadata={"filename": filename, "text_len": len(text)},
        )

        log_ingestion_event(source_id, twin_id, "info", f"YouTube transcript extracted: {len(text)} characters")

        # Direct indexing
        num_chunks = await process_and_index_text(
            source_id,
            twin_id,
            text,
            metadata_override={
                "filename": filename,
                "type": "youtube",
                "video_id": video_id,
                "url": url,
            },
            provider=provider,
            correlation_id=correlation_id,
        )

        # Fetch tenant_id
        tenant_id = None
        try:
            twin_res = supabase.table("twins").select("tenant_id").eq("id", twin_id).single().execute()
            tenant_id = twin_res.data.get("tenant_id") if twin_res.data else None
        except Exception:
            pass

        # Phase 8: Log the action
        AuditLogger.log(
            tenant_id=tenant_id,
            twin_id=twin_id, 
            event_type="KNOWLEDGE_UPDATE", 
            action="SOURCE_INDEXED", 
            metadata={"source_id": source_id, "filename": filename, "type": "youtube", "chunks": num_chunks}
        )

        # Set status to live after successful Pinecone upsert
        supabase.table("sources").update({
            "status": "live",
            "staging_status": "live",
            "chunk_count": num_chunks
        }).eq("id", source_id).execute()

        log_ingestion_event(source_id, twin_id, "info", f"YouTube content indexed: {num_chunks} chunks, status=live")

        live_event_id = start_step(
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="live",
            correlation_id=correlation_id,
            message="Source is live",
        )
        finish_step(
            event_id=live_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="live",
            status="completed",
            correlation_id=correlation_id,
            metadata={"chunks": num_chunks},
        )

        # Enqueue async content extraction to build graph nodes/edges
        try:
            from modules._core.scribe_engine import enqueue_content_extraction_job
            max_chunks = int(os.getenv("CONTENT_EXTRACT_MAX_CHUNKS", "6"))
            enqueue_content_extraction_job(
                twin_id=twin_id,
                source_id=source_id,
                tenant_id=tenant_id,
                source_type="youtube",
                max_chunks=max_chunks
            )
            log_ingestion_event(source_id, twin_id, "info", f"Graph extraction queued (max_chunks={max_chunks})")
        except Exception as e:
            log_ingestion_event(source_id, twin_id, "warning", f"Graph extraction enqueue failed: {e}")
            print(f"[Ingestion] Warning: Failed to enqueue graph extraction for source {source_id}: {e}")

        return num_chunks
    except Exception as e:
        try:
            err = build_error(
                code="YOUTUBE_INDEXING_FAILED",
                message=str(e),
                provider=provider,
                step="indexed",
                correlation_id=correlation_id,
                raw={"video_id": video_id, "url": url},
                exc=e,
            )
            finish_step(
                event_id="",
                source_id=source_id,
                twin_id=twin_id,
                provider=provider,
                step="indexed",
                status="error",
                correlation_id=correlation_id,
                error=err,
            )
        except Exception:
            pass
        print(f"Error processing YouTube content: {e}")
        log_ingestion_event(source_id, twin_id, "error", f"Error processing YouTube content: {e}")
        supabase.table("sources").update({"status": "error", "health_status": "failed"}).eq("id", source_id).execute()
        raise e


async def ingest_podcast_rss(source_id: str, twin_id: str, url: str):
    """
    Ingests the latest episode from a podcast RSS feed.
    """
    try:
        feed = feedparser.parse(url)
        if not feed.entries:
            raise ValueError("No episodes found in RSS feed")

        latest_episode = feed.entries[0]
        audio_url = None
        for enclosure in latest_episode.enclosures:
            if enclosure.type.startswith('audio'):
                audio_url = enclosure.href
                break

        if not audio_url:
            raise ValueError("No audio URL found in the latest episode")

        # Download audio temporarily
        temp_dir = "temp_uploads"
        os.makedirs(temp_dir, exist_ok=True)
        filename = f"{uuid.uuid4()}.mp3"
        file_path = os.path.join(temp_dir, filename)

        import httpx
        async with httpx.AsyncClient() as client:
            response = await client.get(audio_url)
            with open(file_path, "wb") as f:
                f.write(response.content)

        # Transcribe and index (auto-indexed)
        num_chunks = await ingest_source(source_id, twin_id, file_path, f"Podcast: {latest_episode.title}")
        return num_chunks
    except Exception as e:
        print(f"Error ingesting podcast: {e}")
        raise e


async def ingest_x_thread(source_id: str, twin_id: str, url: str, correlation_id: Optional[str] = None):
    """
    Ingests an X (Twitter) thread using multiple fallback strategies.
    """
    tweet_id_match = re.search(r'status/(\d+)', url)
    if not tweet_id_match:
        raise ValueError("Invalid X (Twitter) URL")

    provider = "x"
    tweet_id = tweet_id_match.group(1)
    text = ""
    user = "Unknown"

    # Ensure source row exists
    try:
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": f"X Thread: {tweet_id}",
            "file_size": 0,
            "content_text": "",
            "status": "processing",
            "staging_status": "staged",
            "citation_url": url,
        }).execute()
    except Exception:
        pass

    fetch_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        correlation_id=correlation_id,
        message="Fetching X thread content",
        metadata={"tweet_id": tweet_id, "url": url},
    )

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    # -------------------------------------------------------------
    # Strategy 1: cdn.syndication.twimg.com (legacy, often fails)
    # -------------------------------------------------------------
    try:
        syndication_url = f"https://cdn.syndication.twimg.com/tweet-result?id={tweet_id}&token=0"
        async with httpx.AsyncClient(timeout=15, headers=headers) as client:
            response = await client.get(syndication_url)
            if response.status_code == 200:
                data = response.json()
                text = data.get("text", "")
                user = data.get("user", {}).get("name", "Unknown")
                if text:
                    print(f"[X Thread] Syndication API returned {len(text)} chars")
    except Exception as e:
        print(f"[X Thread] Syndication API failed: {e}")

    # -------------------------------------------------------------
    # Strategy 2: Nitter instances (public Twitter readers)
    # -------------------------------------------------------------
    if not text:
        nitter_instances = [
            "nitter.privacydev.net",
            "nitter.poast.org", 
            "nitter.1d4.us",
        ]
        for instance in nitter_instances:
            try:
                nitter_url = f"https://{instance}/i/status/{tweet_id}"
                async with httpx.AsyncClient(timeout=15, headers=headers, follow_redirects=True) as client:
                    response = await client.get(nitter_url)
                    if response.status_code == 200:
                        page = response.text
                        # Extract tweet content from nitter HTML
                        content_match = re.search(r'<div class="tweet-content[^"]*"[^>]*>(.*?)</div>', page, re.DOTALL)
                        if content_match:
                            raw_text = content_match.group(1)
                            # Clean HTML tags and entities
                            text = re.sub(r'<[^>]+>', ' ', raw_text)
                            text = html_lib.unescape(text)
                            text = re.sub(r'\s+', ' ', text).strip()
                            
                            # Extract username
                            user_match = re.search(r'<a class="fullname"[^>]*>([^<]+)</a>', page)
                            if user_match:
                                user = user_match.group(1).strip()
                            
                            if text:
                                print(f"[X Thread] Nitter {instance} returned {len(text)} chars")
                                break
            except Exception as e:
                print(f"[X Thread] Nitter {instance} failed: {e}")
                continue

    # -------------------------------------------------------------
    # Strategy 3: FxTwitter API (embed generator)
    # -------------------------------------------------------------
    if not text:
        try:
            fx_url = f"https://api.fxtwitter.com/status/{tweet_id}"
            async with httpx.AsyncClient(timeout=15, headers=headers) as client:
                response = await client.get(fx_url)
                if response.status_code == 200:
                    data = response.json()
                    tweet = data.get("tweet", {})
                    text = tweet.get("text", "")
                    user = tweet.get("author", {}).get("name", "Unknown")
                    if text:
                        print(f"[X Thread] FxTwitter API returned {len(text)} chars")
        except Exception as e:
            print(f"[X Thread] FxTwitter API failed: {e}")

    # -------------------------------------------------------------
    # Strategy 4: VxTwitter API (another embed generator)
    # -------------------------------------------------------------
    if not text:
        try:
            vx_url = f"https://api.vxtwitter.com/status/{tweet_id}"
            async with httpx.AsyncClient(timeout=15, headers=headers) as client:
                response = await client.get(vx_url)
                if response.status_code == 200:
                    data = response.json()
                    text = data.get("text", "")
                    user = data.get("user_name", "Unknown")
                    if text:
                        print(f"[X Thread] VxTwitter API returned {len(text)} chars")
        except Exception as e:
            print(f"[X Thread] VxTwitter API failed: {e}")

    # -------------------------------------------------------------
    # Final: Handle result
    # -------------------------------------------------------------
    if not text:
        log_ingestion_event(source_id, twin_id, "error", "All X thread extraction methods failed")
        err = build_error(
            code="X_BLOCKED_OR_UNSUPPORTED",
            message=(
                "Could not extract X thread content. X/Twitter may be blocking requests from this server. "
                "Try copying the tweet text manually and uploading it as a text file."
            ),
            provider=provider,
            step="fetching",
            correlation_id=correlation_id,
            raw={"tweet_id": tweet_id, "url": url},
        )
        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise ValueError(err["message"])

    finish_step(
        event_id=fetch_event_id,
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        status="completed",
        correlation_id=correlation_id,
        metadata={"text_len": len(text)},
    )

    parsed_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="parsed",
        correlation_id=correlation_id,
        message="Persisting extracted X thread text",
        metadata={"tweet_id": tweet_id},
    )

    try:
        content_hash = calculate_content_hash(text)

        # Record source in Supabase - use upsert to be safe
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": f"X Thread: {tweet_id} by {user}",
            "file_size": len(text),
            "content_text": text,
            "content_hash": content_hash,
            "status": "processing",
            "staging_status": "staged",
            "extracted_text_length": len(text)
        }).execute()

        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="completed",
            correlation_id=correlation_id,
            metadata={"text_len": len(text), "user": user},
        )

        # Small verification wait to settle FK
        await asyncio.sleep(1)

        log_ingestion_event(source_id, twin_id, "info", f"X thread extracted: {len(text)} characters")

        # Direct indexing
        num_chunks = await process_and_index_text(source_id, twin_id, text, metadata_override={
            "filename": f"X Thread: {tweet_id} by {user}",
            "type": "x_thread",
            "tweet_id": tweet_id
        }, provider=provider, correlation_id=correlation_id)

        # Fetch tenant_id
        tenant_id = None
        try:
            twin_res = supabase.table("twins").select("tenant_id").eq("id", twin_id).single().execute()
            tenant_id = twin_res.data.get("tenant_id") if twin_res.data else None
        except Exception:
            pass

        AuditLogger.log(
            tenant_id=tenant_id,
            twin_id=twin_id, 
            event_type="KNOWLEDGE_UPDATE", 
            action="SOURCE_INDEXED", 
            metadata={"source_id": source_id, "filename": f"X Thread: {tweet_id} by {user}", "type": "x_thread", "chunks": num_chunks}
        )

        # Set status to live after successful Pinecone upsert
        supabase.table("sources").update({
            "status": "live",
            "staging_status": "live",
            "chunk_count": num_chunks
        }).eq("id", source_id).execute()

        log_ingestion_event(source_id, twin_id, "info", f"X Thread content indexed: {num_chunks} chunks, status=live")

        live_event_id = start_step(
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="live",
            correlation_id=correlation_id,
            message="Source is live",
        )
        finish_step(
            event_id=live_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="live",
            status="completed",
            correlation_id=correlation_id,
            metadata={"chunks": num_chunks},
        )

        # Enqueue async content extraction to build graph nodes/edges
        try:
            from modules._core.scribe_engine import enqueue_content_extraction_job
            max_chunks = int(os.getenv("CONTENT_EXTRACT_MAX_CHUNKS", "6"))
            enqueue_content_extraction_job(
                twin_id=twin_id,
                source_id=source_id,
                tenant_id=tenant_id,
                source_type="twitter",
                max_chunks=max_chunks
            )
            log_ingestion_event(source_id, twin_id, "info", f"Graph extraction queued (max_chunks={max_chunks})")
        except Exception as e:
            log_ingestion_event(source_id, twin_id, "warning", f"Graph extraction enqueue failed: {e}")
            print(f"[Ingestion] Warning: Failed to enqueue graph extraction for source {source_id}: {e}")

        return num_chunks
    except Exception as e:
        try:
            err = build_error(
                code="X_INDEXING_FAILED",
                message=str(e),
                provider=provider,
                step="indexed",
                correlation_id=correlation_id,
                raw={"tweet_id": tweet_id, "url": url},
                exc=e,
            )
            finish_step(
                event_id="",
                source_id=source_id,
                twin_id=twin_id,
                provider=provider,
                step="indexed",
                status="error",
                correlation_id=correlation_id,
                error=err,
            )
        except Exception:
            pass
        print(f"Error processing X thread: {e}")
        log_ingestion_event(source_id, twin_id, "error", f"Error processing X thread: {e}")
        supabase.table("sources").update({"status": "error", "health_status": "failed"}).eq("id", source_id).execute()
        raise e


def _extract_og(soup: BeautifulSoup, prop: str) -> Optional[str]:
    tag = soup.find("meta", attrs={"property": prop})
    if tag and tag.get("content"):
        return str(tag.get("content")).strip()
    return None


def _extract_meta_name(soup: BeautifulSoup, name: str) -> Optional[str]:
    tag = soup.find("meta", attrs={"name": name})
    if tag and tag.get("content"):
        return str(tag.get("content")).strip()
    return None


def _extract_canonical(soup: BeautifulSoup) -> Optional[str]:
    tag = soup.find("link", attrs={"rel": "canonical"})
    if tag and tag.get("href"):
        return str(tag.get("href")).strip()
    return None


def _linkedin_login_wall(html_text: str, final_url: str) -> bool:
    t = (html_text or "").lower()
    u = (final_url or "").lower()
    if "/login" in u or "/checkpoint" in u or "authwall" in u:
        return True
    # Common LinkedIn wall cues
    if "sign in" in t and "linkedin" in t and ("join linkedin" in t or "sign in to see" in t):
        return True
    if "linkedin login" in t or "linkedin: log in" in t:
        return True
    return False


async def ingest_linkedin_open_graph(source_id: str, twin_id: str, url: str, correlation_id: Optional[str] = None) -> int:
    """
    Compliance-first LinkedIn ingestion:
    - Fetches only public OpenGraph/canonical metadata when available.
    - If blocked/login wall, fails with a clear terminal error instructing export/PDF fallback.
    """
    provider = "linkedin"

    # Ensure source row exists
    try:
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": "LinkedIn: queued",
            "file_size": 0,
            "content_text": "",
            "status": "processing",
            "staging_status": "staged",
            "citation_url": url,
        }).execute()
    except Exception as e:
        print(f"[LinkedIn] Warning: Failed to upsert source record: {e}")

    fetch_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        correlation_id=correlation_id,
        message="Fetching LinkedIn OpenGraph metadata",
        metadata={"url": url},
    )

    http_status: Optional[int] = None
    final_url: str = url
    og_title = None
    og_desc = None
    og_image = None
    canonical = None
    page_title = None

    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.8",
        }
        async with httpx.AsyncClient(timeout=20, headers=headers, follow_redirects=True) as client:
            resp = await client.get(url)
            http_status = resp.status_code
            final_url = str(resp.url)
            html_text = resp.text or ""

        if http_status in (401, 403, 429, 999) or _linkedin_login_wall(html_text, final_url):
            # LinkedIn commonly returns HTTP 999 for bot detection.
            raise ValueError("LinkedIn blocked access or requires authentication.")

        soup = BeautifulSoup(html_text, "html.parser")
        og_title = _extract_og(soup, "og:title")
        og_desc = _extract_og(soup, "og:description")
        og_image = _extract_og(soup, "og:image")
        canonical = _extract_canonical(soup) or _extract_og(soup, "og:url")
        page_title = (soup.title.string.strip() if soup.title and soup.title.string else None)

        # Fallback: standard meta tags
        if not og_desc:
            og_desc = _extract_meta_name(soup, "description")

        # Block heuristics: LinkedIn login pages often expose generic titles
        if (og_title and "linkedin login" in og_title.lower()) or (page_title and "linkedin login" in page_title.lower()):
            raise ValueError("LinkedIn returned a login wall page.")

        if not og_title and not og_desc:
            raise ValueError("No public OpenGraph metadata available.")

        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="completed",
            correlation_id=correlation_id,
            metadata={"http_status": http_status, "has_og_title": bool(og_title), "has_og_desc": bool(og_desc)},
        )
    except Exception as e:
        err = build_error(
            code="LINKEDIN_BLOCKED_OR_REQUIRES_AUTH",
            message=(
                "LinkedIn profile content could not be fetched publicly (blocked/login wall). "
                "Upload your LinkedIn profile PDF export or paste profile text for full ingestion."
            ),
            provider=provider,
            step="fetching",
            http_status=http_status,
            correlation_id=correlation_id,
            raw={"url": url, "final_url": final_url},
            exc=e,
        )
        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise ValueError(err["message"])

    parsed_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="parsed",
        correlation_id=correlation_id,
        message="Building minimal profile document from OG fields",
    )
    try:
        title = og_title or page_title or "LinkedIn Profile"
        desc = og_desc or ""
        canon = canonical or final_url or url

        doc = "\n".join([
            "LinkedIn Profile (public metadata)",
            f"Title: {title}",
            f"Description: {desc}",
            f"URL: {canon}",
            f"Image: {og_image or ''}",
        ]).strip()

        content_hash = calculate_content_hash(doc)

        supabase.table("sources").update({
            "filename": f"LinkedIn: {title}"[:240],
            "file_size": len(doc),
            "content_text": doc,
            "content_hash": content_hash,
            "status": "processing",
            "staging_status": "staged",
            "extracted_text_length": len(doc),
            "citation_url": url,
        }).eq("id", source_id).eq("twin_id", twin_id).execute()

        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="completed",
            correlation_id=correlation_id,
            metadata={"text_len": len(doc), "canonical": canon},
        )
    except Exception as e:
        err = build_error(
            code="LINKEDIN_PARSE_FAILED",
            message=f"Failed to build LinkedIn metadata document: {str(e)}",
            provider=provider,
            step="parsed",
            correlation_id=correlation_id,
            raw={"url": url},
            exc=e,
        )
        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    num_chunks = await process_and_index_text(
        source_id,
        twin_id,
        doc,
        metadata_override={
            "filename": f"LinkedIn: {title}"[:240],
            "type": "linkedin_profile_og",
            "url": url,
            "canonical_url": canonical or final_url,
        },
        provider=provider,
        correlation_id=correlation_id,
    )

    supabase.table("sources").update({
        "status": "live",
        "staging_status": "live",
        "chunk_count": num_chunks,
    }).eq("id", source_id).execute()

    live_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="live",
        correlation_id=correlation_id,
        message="Source is live",
    )
    finish_step(
        event_id=live_event_id,
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="live",
        status="completed",
        correlation_id=correlation_id,
        metadata={"chunks": num_chunks},
    )

    return num_chunks


async def ingest_web_url(source_id: str, twin_id: str, url: str, correlation_id: Optional[str] = None) -> int:
    """Ingest a generic web URL by fetching HTML, extracting text, and indexing."""
    provider = "web"

    try:
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": f"Web: {url}"[:240],
            "file_size": 0,
            "content_text": "",
            "status": "processing",
            "staging_status": "staged",
            "citation_url": url,
        }).execute()
    except Exception as e:
        print(f"[Web] Warning: Failed to upsert source record: {e}")

    fetch_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="fetching",
        correlation_id=correlation_id,
        message="Fetching web page",
        metadata={"url": url},
    )

    html_text = ""
    http_status = None
    try:
        async with httpx.AsyncClient(timeout=20, follow_redirects=True) as client:
            resp = await client.get(url)
            http_status = resp.status_code
            resp.raise_for_status()
            html_text = resp.text or ""

        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="completed",
            correlation_id=correlation_id,
            metadata={"http_status": http_status, "bytes": len(html_text)},
        )
    except Exception as e:
        err = build_error(
            code="WEB_FETCH_FAILED",
            message=f"Failed to fetch URL: {str(e)}",
            provider=provider,
            step="fetching",
            http_status=http_status,
            correlation_id=correlation_id,
            raw={"url": url},
            exc=e,
        )
        finish_step(
            event_id=fetch_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="fetching",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    parsed_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="parsed",
        correlation_id=correlation_id,
        message="Extracting text from HTML",
    )
    try:
        soup = BeautifulSoup(html_text, "html.parser")
        title = soup.title.string.strip() if soup.title and soup.title.string else url

        text = soup.get_text(separator="\n", strip=True)
        text = re.sub(r"\n{3,}", "\n\n", text).strip()
        if not text:
            raise ValueError("No text extracted from HTML")

        content_hash = calculate_content_hash(text)
        supabase.table("sources").update({
            "filename": f"Web: {title}"[:240],
            "file_size": len(text),
            "content_text": text,
            "content_hash": content_hash,
            "status": "processing",
            "staging_status": "staged",
            "extracted_text_length": len(text),
            "citation_url": url,
        }).eq("id", source_id).eq("twin_id", twin_id).execute()

        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="completed",
            correlation_id=correlation_id,
            metadata={"text_len": len(text), "title": title},
        )
    except Exception as e:
        err = build_error(
            code="WEB_PARSE_FAILED",
            message=f"Failed to parse page: {str(e)}",
            provider=provider,
            step="parsed",
            correlation_id=correlation_id,
            raw={"url": url},
            exc=e,
        )
        finish_step(
            event_id=parsed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="parsed",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    num_chunks = await process_and_index_text(
        source_id,
        twin_id,
        text,
        metadata_override={"filename": f"Web: {title}"[:240], "type": "web", "url": url},
        provider=provider,
        correlation_id=correlation_id,
    )

    supabase.table("sources").update({
        "status": "live",
        "staging_status": "live",
        "chunk_count": num_chunks,
    }).eq("id", source_id).execute()

    live_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="live",
        correlation_id=correlation_id,
        message="Source is live",
    )
    finish_step(
        event_id=live_event_id,
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="live",
        status="completed",
        correlation_id=correlation_id,
        metadata={"chunks": num_chunks},
    )

    return num_chunks


async def ingest_url_to_source(
    source_id: str,
    twin_id: str,
    url: str,
    provider: Optional[str] = None,
    correlation_id: Optional[str] = None
) -> int:
    """Route a URL ingestion to the correct provider handler using an existing source_id."""
    detected = provider or detect_url_provider(url)
    if detected == "youtube":
        return await ingest_youtube_transcript(source_id, twin_id, url, correlation_id=correlation_id)
    if detected == "x":
        return await ingest_x_thread(source_id, twin_id, url, correlation_id=correlation_id)
    if detected == "podcast":
        return await ingest_podcast_rss(source_id, twin_id, url)
    if detected == "linkedin":
        return await ingest_linkedin_open_graph(source_id, twin_id, url, correlation_id=correlation_id)
    if detected == "web":
        return await ingest_web_url(source_id, twin_id, url, correlation_id=correlation_id)
    raise ValueError(f"Unsupported URL provider: {detected}")


async def transcribe_audio(file_path: str) -> str:
    """
    Transcribes audio using OpenAI Whisper API.
    """
    client = get_openai_client()
    with open(file_path, "rb") as audio_file:
        transcript = client.audio.transcriptions.create(
            model="whisper-1", 
            file=audio_file
        )
    return transcript.text


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
    chunks = []
    for i in range(0, len(text), chunk_size - overlap):
        chunks.append(text[i:i + chunk_size])
    return chunks




async def analyze_chunk_content(text: str) -> dict:
    """
    Analyzes a chunk to generate synthetic questions, category (Fact/Opinion), and tone.
    """
    client = get_openai_client()
    try:
        # Using gpt-4o-mini for better reasoning with JSON output
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": """Analyze the text chunk provided. Return a JSON object with:
                - 'questions': 3 brief questions this text chunk answers.
                - 'category': 'OPINION' if it contains beliefs, values, or personal perspectives. 'FACT' if it is objective information.
                - 'tone': A single word describing the style (e.g., 'Assertive', 'Casual', 'Technical', 'Thoughtful').
                - 'opinion_map': If category is 'OPINION', provide a JSON object with:
                    - 'topic': The main subject of the opinion.
                    - 'stance': A short description of the owner's position.
                    - 'intensity': A score from 1-10 on how strongly this opinion is held.
                  If category is 'FACT', set 'opinion_map' to null."""},
                {"role": "user", "content": text}
            ],
            response_format={ "type": "json_object" }
        )
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        print(f"Error analyzing chunk: {e}")
        return {"questions": [], "category": "FACT", "tone": "Neutral", "opinion_map": None}


async def process_and_index_text(
    source_id: str,
    twin_id: str,
    text: str,
    metadata_override: dict = None,
    provider: str = "unknown",
    correlation_id: Optional[str] = None
):
    # Step: chunked
    chunk_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="chunked",
        correlation_id=correlation_id,
        message="Chunking text",
    )
    try:
        chunks = chunk_text(text)
        finish_step(
            event_id=chunk_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="chunked",
            status="completed",
            correlation_id=correlation_id,
            metadata={"chunks": len(chunks)},
        )
    except Exception as e:
        err = build_error(
            code="CHUNKING_FAILED",
            message=str(e),
            provider=provider,
            step="chunked",
            correlation_id=correlation_id,
            raw={"text_len": len(text) if text else 0},
            exc=e,
        )
        finish_step(
            event_id=chunk_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="chunked",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    # 0. Cleanup existing chunks for this source (idempotency)
    from modules.observability import supabase
    try:
        supabase.table("chunks").delete().eq("source_id", source_id).execute()
    except Exception as e:
        print(f"[Ingestion] Warning: Failed to clean old chunks for source {source_id}: {e}")

    # Step: embedded
    embed_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="embedded",
        correlation_id=correlation_id,
        message="Generating embeddings",
        metadata={"chunks": len(chunks)},
    )

    # Generate embeddings and prepare vectors
    index = get_pinecone_index()
    vectors = []
    db_chunks = []
    
    try:
        for _i, chunk in enumerate(chunks):
            vector_id = str(uuid.uuid4())
            chunk_id = str(uuid.uuid4())  # Supabase primary key

            # Analyze chunk for enrichment
            analysis = await analyze_chunk_content(chunk)
            synth_questions = analysis.get("questions", [])

            # Enriched embedding: include synthetic questions to improve retrieval
            enriched_text = f"CONTENT: {chunk}\nQUESTIONS: {', '.join(synth_questions)}"
            embedding = get_embedding(enriched_text)

            metadata = {
                "source_id": source_id,
                "twin_id": twin_id,
                "chunk_id": chunk_id,  # Link back to DB chunk row
                "text": chunk,  # Keep original text for grounding
                "synthetic_questions": synth_questions,
                "category": analysis.get("category", "FACT"),
                "tone": analysis.get("tone", "Neutral"),
                "is_verified": False  # Explicitly mark regular sources as not verified
            }

            # Add opinion mapping if present
            opinion_map = analysis.get("opinion_map")
            if opinion_map and isinstance(opinion_map, dict):
                metadata["opinion_topic"] = opinion_map.get("topic")
                metadata["opinion_stance"] = opinion_map.get("stance")
                metadata["opinion_intensity"] = opinion_map.get("intensity")

            if metadata_override:
                metadata.update(metadata_override)

            vectors.append({
                "id": vector_id,
                "values": embedding,
                "metadata": metadata
            })
            
            db_chunks.append({
                "id": chunk_id,
                "source_id": source_id,
                "content": chunk,
                "vector_id": vector_id,
                "metadata": metadata
            })

        finish_step(
            event_id=embed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="embedded",
            status="completed",
            correlation_id=correlation_id,
            metadata={"chunks": len(chunks), "vectors": len(vectors)},
        )
    except Exception as e:
        err = build_error(
            code="EMBEDDINGS_FAILED",
            message=str(e),
            provider=provider,
            step="embedded",
            correlation_id=correlation_id,
            raw={"chunks": len(chunks)},
            exc=e,
        )
        finish_step(
            event_id=embed_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="embedded",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    # Step: indexed (Supabase chunks + Pinecone upsert + permissions)
    index_event_id = start_step(
        source_id=source_id,
        twin_id=twin_id,
        provider=provider,
        step="indexed",
        correlation_id=correlation_id,
        message="Persisting chunks and upserting vectors",
        metadata={"vectors": len(vectors)},
    )
    try:
        # Persist chunks to Supabase for citation grounding
        if db_chunks:
            supabase.table("chunks").insert(db_chunks).execute()
            print(f"[Supabase] Persisted {len(db_chunks)} chunks for source_id={source_id}")

        # Upsert vectors to Pinecone (namespace = twin_id for isolation)
        if vectors:
            index.upsert(vectors=vectors, namespace=twin_id)
            print(f"[Pinecone] Upserted {len(vectors)} vectors to namespace={twin_id}")

        # Ensure default group has access to this source (required for retrieval filtering)
        try:
            default_group = await get_default_group(twin_id)
            if default_group and default_group.get("id"):
                await add_content_permission(default_group["id"], "source", source_id, twin_id)
                log_ingestion_event(source_id, twin_id, "info", "Default group permission granted for source")
        except Exception as e:
            log_ingestion_event(source_id, twin_id, "warning", f"Failed to grant default group permission: {e}")
            print(f"[Ingestion] Warning: Failed to grant default group permission for source {source_id}: {e}")

        finish_step(
            event_id=index_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="indexed",
            status="completed",
            correlation_id=correlation_id,
            metadata={"vectors": len(vectors), "chunks": len(db_chunks)},
        )
    except Exception as e:
        err = build_error(
            code="INDEXING_FAILED",
            message=str(e),
            provider=provider,
            step="indexed",
            correlation_id=correlation_id,
            raw={"vectors": len(vectors), "chunks": len(db_chunks)},
            exc=e,
        )
        finish_step(
            event_id=index_event_id,
            source_id=source_id,
            twin_id=twin_id,
            provider=provider,
            step="indexed",
            status="error",
            correlation_id=correlation_id,
            error=err,
        )
        raise

    return len(vectors)


def _infer_source_type(filename: str) -> str:
    name = (filename or "").lower()
    if "youtube" in name or "youtu.be" in name:
        return "youtube"
    if "podcast" in name or "rss" in name or "anchor.fm" in name or "podbean" in name:
        return "podcast"
    if "x thread" in name or "twitter.com" in name or "x.com" in name:
        return "twitter"
    if name.endswith(".pdf"):
        return "pdf"
    if name.endswith(".docx"):
        return "docx"
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.startswith("http://") or name.startswith("https://"):
        return "url"
    return "ingested_content"


async def ingest_source(source_id: str, twin_id: str, file_path: str, filename: str = None):
    """Ingest a file - extracts text and indexes to Pinecone.
    
    Args:
        source_id: Unique identifier for this source
        twin_id: Owner twin ID
        file_path: Path to the file to ingest
        filename: Optional display name for the file
    """
    # 0. Check for existing sources with same name to handle "update"
    has_duplicate = False
    if filename:
        existing = supabase.table("sources").select("id").eq("twin_id", twin_id).eq("filename", filename).execute()
        if existing.data:
            print(f"File {filename} already exists. Updating source(s)...")
            has_duplicate = True
            # Delete ALL old versions first to keep knowledge clean
            for record in existing.data:
                old_source_id = record["id"]
                await delete_source(old_source_id, twin_id)
            # We keep the new source_id for the new record

    # 0.1 Record source in Supabase FIRST (before logging - prevents FK error).
    # Use upsert to support queued sources (row may already exist).
    if filename:
        file_size = os.path.getsize(file_path)
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            "filename": filename,
            "file_size": file_size,
            "status": "processing"
        }).execute()

    # Log after source record exists (to satisfy FK constraint)
    if has_duplicate:
        log_ingestion_event(source_id, twin_id, "info", "Duplicate filename detected, removed old sources")

    # 1. Extract text (PDF, Docx, Excel, or Audio)
    if file_path.endswith('.pdf'):
        text = extract_text_from_pdf(file_path)
    elif file_path.endswith('.docx'):
        text = extract_text_from_docx(file_path)
    elif file_path.endswith('.xlsx'):
        text = extract_text_from_excel(file_path)
    elif file_path.endswith(('.mp3', '.wav', '.m4a', '.webm')):
        text = await transcribe_audio(file_path)
    else:
        # Generic text extraction for other types if needed, or error
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()

    # Extract text and run health checks before indexing
    content_hash = calculate_content_hash(text)

    # Update source with extracted text and processing status
    update_data = {
        "content_text": text,
        "content_hash": content_hash,
        "status": "processing",
        "staging_status": "staged",  # Auto-indexes, goes to 'live' after indexing
        "extracted_text_length": len(text)
    }

    if filename:
        supabase.table("sources").update(update_data).eq("id", source_id).execute()
    else:
        # If no filename, we need to create the source record (upsert for safety).
        supabase.table("sources").upsert({
            "id": source_id,
            "twin_id": twin_id,
            **update_data
        }).execute()

    log_ingestion_event(source_id, twin_id, "info", f"Text extracted: {len(text)} characters")

    tenant_id = None
    try:
        twin_res = supabase.table("twins").select("tenant_id").eq("id", twin_id).single().execute()
        tenant_id = twin_res.data.get("tenant_id") if twin_res.data else None
    except Exception:
        pass

    # Phase 9: Log the action
    AuditLogger.log(
        tenant_id=tenant_id,
        twin_id=twin_id,
        event_type="KNOWLEDGE_UPDATE",
        action="SOURCE_EXTRACTED",
        metadata={"source_id": source_id, "filename": filename or "unknown", "type": "file_upload"}
    )

    # Run health checks
    health_result = run_all_health_checks(
        source_id, 
        twin_id, 
        text,
        source_data={
            "filename": filename or "unknown",
            "twin_id": twin_id
        }
    )

    # Update source health status
    supabase.table("sources").update({
        "health_status": health_result["overall_status"]
    }).eq("id", source_id).execute()

    log_ingestion_event(source_id, twin_id, "info", f"Health checks completed: {health_result['overall_status']}")

    log_ingestion_event(source_id, twin_id, "info", "Auto-indexing enabled")
    num_chunks = await process_and_index_text(source_id, twin_id, text, metadata_override={
        "filename": filename or "unknown",
        "type": "file"
    })
    
    # Set status to live after successful Pinecone upsert
    supabase.table("sources").update({
        "status": "live",
        "staging_status": "live",
        "chunk_count": num_chunks
    }).eq("id", source_id).execute()
    
    log_ingestion_event(source_id, twin_id, "info", f"Auto-indexed: {num_chunks} chunks, status=live")

    # Enqueue async content extraction to build graph nodes/edges
    try:
        from modules._core.scribe_engine import enqueue_content_extraction_job
        max_chunks = int(os.getenv("CONTENT_EXTRACT_MAX_CHUNKS", "6"))
        source_type = _infer_source_type(filename or "")
        enqueue_content_extraction_job(
            twin_id=twin_id,
            source_id=source_id,
            tenant_id=tenant_id,
            source_type=source_type,
            max_chunks=max_chunks
        )
        log_ingestion_event(source_id, twin_id, "info", f"Graph extraction queued (max_chunks={max_chunks})")
    except Exception as e:
        log_ingestion_event(source_id, twin_id, "warning", f"Graph extraction enqueue failed: {e}")
        print(f"[Ingestion] Warning: Failed to enqueue graph extraction for source {source_id}: {e}")
    
    # Audit log
    AuditLogger.log(
        tenant_id=tenant_id,
        twin_id=twin_id,
        event_type="KNOWLEDGE_UPDATE",
        action="SOURCE_INDEXED",
        metadata={
            "source_id": source_id, 
            "filename": filename or "unknown", 
            "type": "file",
            "chunks": num_chunks
        }
    )
    
    return num_chunks


async def delete_source(source_id: str, twin_id: str):
    """
    Deletes a source from Supabase and its associated vectors from Pinecone.
    """
    # 1. Delete from Pinecone
    index = get_pinecone_index()
    try:
        # Note: Delete by filter requires metadata indexing enabled or serverless index
        index.delete(
            filter={
                "source_id": {"$eq": source_id}
            },
            namespace=twin_id
        )
    except Exception as e:
        print(f"Error deleting from Pinecone: {e}")
        # Continue to delete from Supabase even if Pinecone fails (maybe it was already gone)

    # 2. Delete from Supabase
    supabase.table("sources").delete().eq("id", source_id).eq("twin_id", twin_id).execute()

    tenant_id = None
    try:
        twin_res = supabase.table("twins").select("tenant_id").eq("id", twin_id).single().execute()
        tenant_id = twin_res.data.get("tenant_id") if twin_res.data else None
    except Exception:
        pass

    # Phase 9: Log the action
    AuditLogger.log(
        tenant_id=tenant_id,
        twin_id=twin_id,
        event_type="KNOWLEDGE_UPDATE",
        action="SOURCE_DELETED",
        metadata={"source_id": source_id}
    )

    return True


async def bulk_update_source_metadata(source_ids: List[str], metadata_updates: Dict[str, Any]):
    """
    Bulk update metadata for sources.
    
    Args:
        source_ids: List of source UUIDs
        metadata_updates: Dict with fields to update (access_group, publish_date, author, citation_url, visibility)
    """
    allowed_fields = ["publish_date", "author", "citation_url"]
    update_data = {k: v for k, v in metadata_updates.items() if k in allowed_fields}

    if not update_data:
        return

    for source_id in source_ids:
        try:
            supabase.table("sources").update(update_data).eq("id", source_id).execute()
        except Exception as e:
            print(f"Error updating source {source_id}: {e}")


# Wrapper functions for router endpoints (create source_id and call actual functions)


async def ingest_youtube_transcript_wrapper(twin_id: str, url: str) -> str:
    """Wrapper that creates source_id and calls ingest_youtube_transcript"""
    source_id = str(uuid.uuid4())
    await ingest_youtube_transcript(source_id, twin_id, url)
    return source_id


async def ingest_podcast_transcript(twin_id: str, url: str) -> str:
    """Wrapper that creates source_id and calls ingest_podcast_rss"""
    source_id = str(uuid.uuid4())
    await ingest_podcast_rss(source_id, twin_id, url)
    return source_id


async def ingest_x_thread_wrapper(twin_id: str, url: str) -> str:
    """Wrapper that creates source_id and calls ingest_x_thread"""
    source_id = str(uuid.uuid4())
    await ingest_x_thread(source_id, twin_id, url)
    return source_id


async def ingest_file(twin_id: str, file) -> str:
    """Wrapper that saves uploaded file and calls ingest_source.
    
    Args:
        twin_id: Owner twin ID
        file: Uploaded file object
    """

    source_id = str(uuid.uuid4())
    temp_dir = "temp_uploads"
    os.makedirs(temp_dir, exist_ok=True)

    # Save uploaded file temporarily
    file_extension = os.path.splitext(file.filename)[1] if file.filename else ""
    temp_filename = f"{source_id}{file_extension}"
    file_path = os.path.join(temp_dir, temp_filename)

    try:
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        # Call ingest_source with the file path
        await ingest_source(source_id, twin_id, file_path, file.filename)

        # Clean up temp file
        if os.path.exists(file_path):
            os.remove(file_path)

        return source_id
    except Exception as e:
        # Clean up temp file on error
        if os.path.exists(file_path):
            os.remove(file_path)
        raise e


async def ingest_url(twin_id: str, url: str) -> str:
    """Wrapper that creates a new source_id and ingests a URL."""
    source_id = str(uuid.uuid4())
    await ingest_url_to_source(source_id=source_id, twin_id=twin_id, url=url)
    return source_id
